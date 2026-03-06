from flask import Flask, render_template, request, jsonify, Response
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from catboost import CatBoostRegressor
import pandas as pd
import os
import re
import requests
import json
import csv
from datetime import datetime
from typing import List, Dict
try:
    from duckduckgo_search import DDGS
    WEB_SEARCH_AVAILABLE = True
except ImportError:
    WEB_SEARCH_AVAILABLE = False
    print("[警告] duckduckgo-search 未安装，联网搜索功能不可用")

app = Flask(__name__)

# --- 接口限流（防止滥用）---
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per hour"],
    storage_uri="memory://",
)

# --- 健康检查（用于排查“网络错误”到底是前端还是后端问题）---
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

# --- 1. 核心辅助函数：把邮编翻译成大区 (必须和训练脚本一致) ---
def get_zone_from_postcode(val):
    # 使用正则表达式提取 4 位数字
    val_str = str(val)
    digits = re.findall(r'\b\d{4}\b', val_str)

    if not digits:
        return 'Unknown_Zone'

    postcode = int(digits[0])

    # 澳洲大区划分规则 (必须与 train_advanced.py 完全一致)
    if 2000 <= postcode <= 2234: return 'NSW_Metro'
    if 2500 <= postcode <= 2530: return 'NSW_Regional'
    if 2600 <= postcode <= 2620: return 'ACT_Metro'
    if 3000 <= postcode <= 3207: return 'VIC_Metro'
    if 4000 <= postcode <= 4179: return 'QLD_Metro'
    if 6000 <= postcode <= 6199: return 'WA_Metro'

    return 'Other_Regional'


# 验证澳洲邮编是否有效（澳洲邮编范围 0200-9999）
VALID_POSTCODE_RANGES = [
    (2000, 2234, 'NSW Metro (悉尼)'),
    (2235, 2499, 'NSW 其他区域'),
    (2500, 2530, 'NSW Regional (卧龙岗)'),
    (2531, 2599, 'NSW 南海岸'),
    (2600, 2620, 'ACT Metro (堪培拉)'),
    (2621, 2899, 'NSW/ACT 其他区域'),
    (2900, 2999, 'ACT 其他区域'),
    (3000, 3207, 'VIC Metro (墨尔本)'),
    (3208, 3999, 'VIC 其他区域'),
    (4000, 4179, 'QLD Metro (布里斯班)'),
    (4180, 4999, 'QLD 其他区域'),
    (5000, 5199, 'SA Metro (阿德莱德)'),
    (5200, 5999, 'SA 其他区域'),
    (6000, 6199, 'WA Metro (珀斯)'),
    (6200, 6999, 'WA 其他区域'),
    (7000, 7999, 'TAS (塔斯马尼亚)'),
    (800, 899, 'NT (北领地)'),
]

# 覆盖的主要区域（有训练数据支持的区域）
COVERED_ZONES = ['NSW_Metro', 'NSW_Regional', 'ACT_Metro', 'VIC_Metro', 'QLD_Metro', 'WA_Metro', 'Other_Regional']

def validate_postcode(val):
    """验证邮编是否为有效的澳洲邮编，返回 (is_valid, region_name)"""
    val_str = str(val)
    digits = re.findall(r'\b\d{3,4}\b', val_str)
    if not digits:
        return False, '无法识别的邮编格式'
    postcode = int(digits[0])
    if postcode < 200 or postcode > 9999:
        return False, f'邮编 {postcode} 不在澳洲邮编范围内 (0200-9999)'
    for low, high, name in VALID_POSTCODE_RANGES:
        if low <= postcode <= high:
            return True, name
    return True, '其他区域'


# --- 2. 启动时加载模型和训练数据（用于置信度计算）---
MODEL_PATH = "telecom_model.cbm"
TRAINING_DATA_PATH = "training_mvp_v5.csv"

if not os.path.exists(MODEL_PATH):
    print(f"[错误] 严重错误：找不到模型文件 {MODEL_PATH}！")
    print("请先运行 python trainadvance.py 生成新模型！")
    exit()

print("正在加载增强版 AI 模型...")
model = CatBoostRegressor()
model.load_model(MODEL_PATH)
print("[成功] 模型加载完毕")

# 加载训练数据用于置信度计算
print("正在加载训练数据用于置信度分析...")
try:
    df_training = pd.read_csv(TRAINING_DATA_PATH)
    # 统一列名
    col_map = {
        'operator_manual': 'operator',
        'product_type_rule': 'product_type',
        'region_rule': 'region',
        'MRC (ex)': 'price',
        'mrc': 'price'
    }
    df_training = df_training.rename(columns=col_map)
    # 添加zone特征（与训练脚本一致）
    df_training['zone'] = df_training['region'].apply(get_zone_from_postcode)
    print(f"[成功] 训练数据加载完毕，共 {len(df_training)} 条记录")
except Exception as e:
    print(f"[警告] 无法加载训练数据: {e}，置信度计算将使用简化算法")
    df_training = None

print("服务启动中...")

# --- AI 模型配置（支持多个免费模型）---
# 可选模型：lmstudio, gemini, qwen, deepseek, ollama
AI_MODEL_TYPE = os.getenv('AI_MODEL_TYPE', 'ollama').lower()  # 默认使用 Ollama（本地免费）

# LM Studio 配置（本地运行，完全免费）
LMSTUDIO_BASE_URL = os.getenv('LMSTUDIO_BASE_URL', 'http://localhost:1234')
LMSTUDIO_MODEL = os.getenv('LMSTUDIO_MODEL', '')  # 使用 LM Studio 中加载的模型

# Ollama 配置（本地运行，完全免费）
OLLAMA_BASE_URL = os.getenv('OLLAMA_BASE_URL', 'http://localhost:11434')
OLLAMA_MODEL = os.getenv('OLLAMA_MODEL', 'qwen2.5:7b')  # 推荐模型：qwen2.5:7b, llama3.2, mistral

# Google Gemini 配置（免费额度大）
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')  # 请通过环境变量配置 API Key
GEMINI_MODEL = os.getenv('GEMINI_MODEL', 'gemini-2.0-flash')  # 使用最新的 gemini-2.0-flash 模型
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

# 通义千问配置（国内访问快，有免费额度）
QWEN_API_KEY = os.getenv('QWEN_API_KEY', '')
QWEN_API_URL = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"

# DeepSeek 配置（备用）
DEEPSEEK_API_KEY = os.getenv('DEEPSEEK_API_KEY', '')
DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"

# --- 加载公司知识库文件（company_knowledge.txt）---
COMPANY_KNOWLEDGE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "company_knowledge.txt")

# 读取失败时使用的默认简介（确保程序不会崩溃）
_DEFAULT_COMPANY_CONTEXT = (
    "澳洲联通 (Australia Unicom) 是澳大利亚企业级网络服务提供商，"
    "提供企业光纤专线 (DIA)、SD-WAN、数据中心互联等服务。"
    "客服热线：1300-UNICOM，24小时中英双语服务，99.99% SLA 保障。"
)


def load_company_knowledge(filepath: str) -> str:
    """
    从 company_knowledge.txt 读取公司内部知识库。

    - 使用 UTF-8 编码打开，兼容中英文混合内容
    - 读取失败时打印错误日志并返回默认简介，不会导致程序崩溃

    Returns:
        str: 知识库文本内容
    """
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if not content:
            print(f"[警告] 知识库文件 {filepath} 为空，使用默认简介")
            return _DEFAULT_COMPANY_CONTEXT
        print(f"[成功] 公司知识库加载完成: {filepath} ({len(content)} 字符)")
        return content
    except FileNotFoundError:
        print(f"[警告] 找不到知识库文件: {filepath}，使用默认简介")
        return _DEFAULT_COMPANY_CONTEXT
    except UnicodeDecodeError as e:
        print(f"[错误] 知识库文件编码异常 (需要 UTF-8): {e}，使用默认简介")
        return _DEFAULT_COMPANY_CONTEXT
    except OSError as e:
        print(f"[错误] 读取知识库文件失败: {e}，使用默认简介")
        return _DEFAULT_COMPANY_CONTEXT


# 启动时加载公司知识库
company_context = load_company_knowledge(COMPANY_KNOWLEDGE_PATH)


# --- 构建训练数据知识库（用于 RAG 数据分析）---
def build_training_data_summary(df: pd.DataFrame) -> str:
    """基于训练数据构建统计摘要，作为 RAG 数据源补充"""
    if df is None or len(df) == 0:
        return ""
    
    parts: list[str] = []
    
    # 数据概览
    parts.append(f"## 训练数据统计摘要")
    parts.append(f"- 总数据量: {len(df)} 条记录")
    
    # 运营商分布
    operator_stats = df['operator'].value_counts()
    parts.append(f"\n## 运营商分布")
    for operator, count in operator_stats.items():
        parts.append(f"- {operator}: {count} 条 ({count / len(df) * 100:.1f}%)")
    
    # 产品类型分布
    product_stats = df['product_type'].value_counts()
    parts.append(f"\n## 产品类型分布")
    for product, count in product_stats.items():
        parts.append(f"- {product}: {count} 条 ({count / len(df) * 100:.1f}%)")
    
    # 地区分布
    zone_stats = df['zone'].value_counts()
    parts.append(f"\n## 地区分布")
    for zone, count in zone_stats.head(10).items():
        parts.append(f"- {zone}: {count} 条 ({count / len(df) * 100:.1f}%)")
    
    # 价格统计
    if 'price' in df.columns:
        ps = df['price'].describe()
        parts.append(f"\n## 价格统计（澳元/月）")
        parts.append(f"- 最低: ${ps['min']:.2f}  最高: ${ps['max']:.2f}")
        parts.append(f"- 平均: ${ps['mean']:.2f}  中位数: ${ps['50%']:.2f}")
    
    # 带宽统计
    if 'bandwidth_mbps' in df.columns:
        bs = df['bandwidth_mbps'].describe()
        parts.append(f"\n## 带宽统计（Mbps）")
        parts.append(f"- 范围: {bs['min']:.0f} ~ {bs['max']:.0f} Mbps")
        parts.append(f"- 平均: {bs['mean']:.0f} Mbps  中位数: {bs['50%']:.0f} Mbps")
    
    # 典型案例
    parts.append(f"\n## 典型价格案例")
    for operator in df['operator'].value_counts().head(5).index:
        op_data = df[df['operator'] == operator]
        for pt in op_data['product_type'].value_counts().head(2).index:
            pt_data = op_data[op_data['product_type'] == pt]
            if len(pt_data) > 0:
                parts.append(
                    f"- {operator} {pt}: 均价 ${pt_data['price'].mean():.2f}/月, "
                    f"均带宽 {pt_data['bandwidth_mbps'].mean():.0f} Mbps "
                    f"({len(pt_data)} 条)"
                )
    
    return "\n".join(parts)


# 构建训练数据摘要
training_data_summary = ""
if df_training is not None:
    training_data_summary = build_training_data_summary(df_training)
    print(f"[成功] 训练数据摘要构建完成，共 {len(training_data_summary)} 字符")


# --- 3. 访问主页 ---
@app.route('/')
def home():
    return render_template('homepage.html')

@app.route('/pricing')
def pricing():
    return render_template('index.html')

# --- 文章页面路由 ---
@app.route('/article/innovation-leadership')
def article_innovation():
    return render_template('article_innovation.html')

@app.route('/article/sme-communication')
def article_sme():
    return render_template('article_sme.html')

@app.route('/article/network-coverage')
def article_coverage():
    return render_template('article_coverage.html')


# --- 项目监察功能 ---
PROJECT_DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "projects.json")

# 项目任务点分类：9 大类，部分含小类（可点击展开）
DEFAULT_STAGES = [
    {"id": 1, "name": "需求确认", "name_en": "Requirement Confirmation", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": []},
    {"id": 2, "name": "方案设计", "name_en": "Solution Design", "sub_categories": ["询价", "解决方案", "线上核查"], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [1]},
    {"id": 3, "name": "合同签署", "name_en": "Contract Signing", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [2]},
    {"id": 4, "name": "采购流程", "name_en": "Procurement", "sub_categories": ["线上转定", "采购需求", "采购合同签署"], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [3]},
    {"id": 5, "name": "下单", "name_en": "Place Order", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [4]},
    {"id": 6, "name": "现场勘察", "name_en": "Site Survey", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [5]},
    {"id": 7, "name": "工程施工", "name_en": "Installation", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [6]},
    {"id": 8, "name": "系统测试", "name_en": "System Testing", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [7]},
    {"id": 9, "name": "项目交付", "name_en": "Project Delivery", "sub_categories": [], "status": "pending", "progress": 0, "notes": "", "assignee": "", "start_date": "", "due_date": "", "started_at": "", "completed_at": "", "blocked_by": [8]},
]

# 上传文件目录
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)

# 通知数据路径
NOTIFICATIONS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "notifications.json")

ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'zip'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def migrate_project(project):
    """向后兼容：对旧格式数据自动补全新字段"""
    # 项目级别新字段
    defaults = {
        'priority': 'medium',
        'team': [],
        'due_date': '',
        'comments': [],
        'activity_log': [],
        'attachments': [],
        'pricing_ref': '',
        'tags': [],
    }
    for key, val in defaults.items():
        if key not in project:
            project[key] = val

    # 阶段级别新字段
    stage_defaults = {
        'assignee': '',
        'start_date': '',
        'due_date': '',
        'started_at': '',
        'completed_at': '',
        'blocked_by': [],
    }
    # 默认依赖链（9 步）
    default_blocked_by = {1: [], 2: [1], 3: [2], 4: [3], 5: [4], 6: [5], 7: [6], 8: [7], 9: [8]}
    stages = project.get('stages', [])
    # 若为旧版 7 步，迁移到 9 步
    if len(stages) == 7 and not any(s.get('name') == '采购流程' for s in stages):
        stage_map = {1: 1, 2: 2, 3: 3, 4: 6, 5: 7, 6: 8, 7: 9}  # 旧 id -> 新 id
        new_stages = []
        for s in stages:
            new_id = stage_map.get(s.get('id'), s.get('id'))
            default_stage = next((d for d in DEFAULT_STAGES if d['id'] == new_id), {})
            new_stages.append({
                'id': new_id, 'name': default_stage.get('name', s.get('name')),
                'name_en': default_stage.get('name_en', s.get('name_en', '')),
                'sub_categories': default_stage.get('sub_categories', []),
                'status': s.get('status', 'pending'), 'progress': s.get('progress', 0),
                'notes': s.get('notes', ''), 'assignee': s.get('assignee', ''),
                'start_date': s.get('start_date', ''), 'due_date': s.get('due_date', ''),
                'started_at': s.get('started_at', ''), 'completed_at': s.get('completed_at', ''),
                'blocked_by': default_blocked_by.get(new_id, []),
            })
        # 插入 采购流程(4)、下单(5)
        proc = {k: v for k, v in DEFAULT_STAGES[3].items() if k not in ('status', 'progress', 'notes', 'assignee', 'start_date', 'due_date', 'started_at', 'completed_at')}
        proc.update({'status': 'pending', 'progress': 0, 'notes': '', 'assignee': '', 'start_date': '', 'due_date': '', 'started_at': '', 'completed_at': '', 'blocked_by': [3]})
        order_stage = {k: v for k, v in DEFAULT_STAGES[4].items() if k not in ('status', 'progress', 'notes', 'assignee', 'start_date', 'due_date', 'started_at', 'completed_at')}
        order_stage.update({'status': 'pending', 'progress': 0, 'notes': '', 'assignee': '', 'start_date': '', 'due_date': '', 'started_at': '', 'completed_at': '', 'blocked_by': [4]})
        project['stages'] = new_stages[:3] + [proc, order_stage] + new_stages[3:]
        stages = project['stages']
    for stage in stages:
        for key, val in stage_defaults.items():
            if key not in stage:
                if key == 'blocked_by':
                    stage[key] = default_blocked_by.get(stage.get('id', 0), [])
                else:
                    stage[key] = val
        if 'sub_categories' not in stage:
            default_stage = next((d for d in DEFAULT_STAGES if d['id'] == stage.get('id')), {})
            stage['sub_categories'] = default_stage.get('sub_categories', [])
    return project


def load_projects():
    """从 JSON 文件加载项目列表（自动迁移旧格式）"""
    if not os.path.exists(PROJECT_DATA_PATH):
        return []
    try:
        with open(PROJECT_DATA_PATH, 'r', encoding='utf-8') as f:
            projects = json.load(f)
        # 自动迁移旧数据
        migrated = False
        for p in projects:
            old_stage_count = len(p.get('stages', []))
            migrate_project(p)
            if len(p.get('stages', [])) != old_stage_count:
                migrated = True
        if migrated:
            save_projects(projects)
            print("[项目监察] 已自动迁移旧数据格式")
        return projects
    except Exception as e:
        print(f"[项目监察] 加载失败: {e}")
        return []


def save_projects(projects):
    """保存项目列表到 JSON 文件"""
    try:
        with open(PROJECT_DATA_PATH, 'w', encoding='utf-8') as f:
            json.dump(projects, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[项目监察] 保存失败: {e}")


@app.route('/projects')
def project_monitor():
    return render_template('project_monitor.html')


@app.route('/projects/analytics')
def project_analytics():
    return render_template('project_analytics.html')


@app.route('/projects/<project_id>')
def project_detail(project_id):
    return render_template('project_detail.html', project_id=project_id)


@app.route('/share/project/<project_id>')
def project_share_view(project_id):
    """公开只读分享页面"""
    projects = load_projects()
    project = next((p for p in projects if p.get('id') == project_id), None)
    if not project:
        return '<h2 style="font-family:sans-serif;text-align:center;margin-top:100px;color:#666">项目不存在或已被删除</h2>', 404
    return render_template('project_share.html', project=project)


@app.route('/api/projects/<project_id>/share/email', methods=['GET'])
def api_project_share_email(project_id):
    """生成可用于邮件的独立 HTML 分享页"""
    projects = load_projects()
    project = next((p for p in projects if p.get('id') == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'}), 404
    html = render_template('project_share_email.html', project=project,
                           generated_at=datetime.now().strftime('%Y-%m-%d %H:%M'))
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/projects/<project_id>/share/eml', methods=['GET'])
def api_project_share_eml(project_id):
    """生成 .eml 文件——双击后在 Outlook/Apple Mail 中以 HTML 格式打开撰写窗口"""
    projects_list = load_projects()
    project = next((p for p in projects_list if p.get('id') == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'}), 404

    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')
    html_body = render_template('project_share_email.html', project=project,
                                generated_at=generated_at)

    # 纯文本备用正文
    status_map = {'active': '进行中', 'completed': '已完成', 'cancelled': '已取消'}
    prio_map   = {'high': '高优先级', 'medium': '中优先级', 'low': '低优先级'}
    stage_icon = {'green': '✅', 'yellow': '⚠️', 'red': '🔴', 'pending': '⏳'}
    lines = [
        f"项目名称：{project.get('name', '')}",
        f"客　　户：{project.get('client', '未指定')}",
        f"产品类型：{project.get('product_type', '-')}",
        f"优先级　：{prio_map.get(project.get('priority', 'medium'), '中优先级')}",
        f"项目状态：{status_map.get(project.get('status', 'active'), project.get('status', ''))}",
        f"总体进度：{project.get('overall_progress', 0)}%",
        "", "── 各阶段进度 ──",
    ]
    for s in project.get('stages', []):
        icon = stage_icon.get(s.get('status', 'pending'), '⏳')
        lines.append(f"  {icon} {s.get('name', '')}：{s.get('progress', 0)}%")
    lines += ["", f"生成时间：{generated_at}", "此邮件由澳洲联通项目管理系统自动生成。"]
    plain_body = "\r\n".join(lines)

    subject = f"【澳洲联通】项目进度报告 - {project.get('name', '')}"

    msg = MIMEMultipart('alternative')
    msg['Subject'] = Header(subject, 'utf-8')
    msg['X-Unsent'] = '1'          # 关键：告知邮件客户端以草稿/撰写模式打开
    msg['MIME-Version'] = '1.0'

    msg.attach(MIMEText(plain_body, 'plain', 'utf-8'))
    msg.attach(MIMEText(html_body,  'html',  'utf-8'))

    filename = f"项目分享_{project.get('name', project_id)}_{datetime.now().strftime('%Y%m%d')}.eml"
    return Response(
        msg.as_string(),
        mimetype='message/rfc822',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{filename}'}
    )


def add_activity_log(project, action, detail):
    """向项目添加一条操作日志"""
    if 'activity_log' not in project:
        project['activity_log'] = []
    project['activity_log'].insert(0, {
        'action': action,
        'detail': detail,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    # 最多保留200条日志
    project['activity_log'] = project['activity_log'][:200]


def add_notification(project_id, project_name, ntype, message):
    """添加一条通知"""
    try:
        notifications = []
        if os.path.exists(NOTIFICATIONS_PATH):
            with open(NOTIFICATIONS_PATH, 'r', encoding='utf-8') as f:
                notifications = json.load(f)
        notifications.insert(0, {
            'id': str(int(datetime.now().timestamp() * 1000)),
            'project_id': project_id,
            'project_name': project_name,
            'type': ntype,  # warning, critical, complete, info
            'message': message,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'read': False
        })
        notifications = notifications[:100]  # 最多100条
        with open(NOTIFICATIONS_PATH, 'w', encoding='utf-8') as f:
            json.dump(notifications, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[通知] 保存失败: {e}")


@app.route('/api/heartbeat', methods=['GET', 'POST'])
def api_heartbeat():
    """前端心跳检测"""
    return jsonify({'status': 'ok'})


@app.route('/api/projects', methods=['GET'])
def api_get_projects():
    """获取所有项目"""
    projects = load_projects()
    return jsonify({'status': 'success', 'projects': projects})


@app.route('/api/projects/stats', methods=['GET'])
def api_get_project_stats():
    """获取项目统计数据（用于分析仪表盘）"""
    projects = load_projects()
    total = len(projects)
    active = len([p for p in projects if p.get('status') == 'active'])
    completed = len([p for p in projects if p.get('status') == 'completed'])
    cancelled = len([p for p in projects if p.get('status') == 'cancelled'])

    # 阶段问题热力图数据
    stage_issues = {}
    for p in projects:
        for s in p.get('stages', []):
            sid = s.get('id', 0)
            sname = s.get('name', '')
            if sid not in stage_issues:
                stage_issues[sid] = {'name': sname, 'yellow': 0, 'red': 0, 'green': 0, 'pending': 0}
            st = s.get('status', 'pending')
            if st in stage_issues[sid]:
                stage_issues[sid][st] += 1

    # 按产品类型分布
    product_dist = {}
    for p in projects:
        pt = p.get('product_type', 'Unknown')
        product_dist[pt] = product_dist.get(pt, 0) + 1

    # 各项目进度
    progress_list = [{'name': p.get('name', ''), 'progress': p.get('overall_progress', 0), 'status': p.get('status', '')} for p in projects]

    # 阶段平均耗时（天）
    stage_durations = {}
    for p in projects:
        for s in p.get('stages', []):
            sid = s.get('id', 0)
            started = s.get('started_at', '')
            completed_at = s.get('completed_at', '')
            if started and completed_at:
                try:
                    d1 = datetime.strptime(started, '%Y-%m-%d %H:%M:%S')
                    d2 = datetime.strptime(completed_at, '%Y-%m-%d %H:%M:%S')
                    days = (d2 - d1).days
                    if sid not in stage_durations:
                        stage_durations[sid] = {'name': s.get('name', ''), 'days': [], 'avg': 0}
                    stage_durations[sid]['days'].append(max(days, 0))
                except:
                    pass
    for sid in stage_durations:
        d = stage_durations[sid]['days']
        stage_durations[sid]['avg'] = round(sum(d) / len(d), 1) if d else 0

    # 月度完成趋势
    monthly = {}
    for p in projects:
        if p.get('status') == 'completed' and p.get('updated_at'):
            month = p['updated_at'][:7]  # YYYY-MM
            monthly[month] = monthly.get(month, 0) + 1

    return jsonify({
        'status': 'success',
        'stats': {
            'total': total, 'active': active, 'completed': completed, 'cancelled': cancelled,
            'stage_issues': stage_issues,
            'product_distribution': product_dist,
            'progress_list': progress_list,
            'stage_durations': stage_durations,
            'monthly_completions': monthly
        }
    })


@app.route('/api/projects/overdue', methods=['GET'])
def api_get_overdue():
    """获取所有超期项目/阶段"""
    projects = load_projects()
    now = datetime.now()
    overdue = []
    for p in projects:
        if p.get('status') != 'active':
            continue
        # 项目级别超期
        if p.get('due_date'):
            try:
                dd = datetime.strptime(p['due_date'], '%Y-%m-%d')
                diff = (dd - now).days
                if diff < 0:
                    overdue.append({'project_id': p['id'], 'project_name': p['name'], 'type': 'project', 'days_overdue': abs(diff)})
            except:
                pass
        # 阶段级别超期
        for s in p.get('stages', []):
            if s.get('status') == 'pending' or s.get('progress', 0) >= 100:
                continue
            if s.get('due_date'):
                try:
                    dd = datetime.strptime(s['due_date'], '%Y-%m-%d')
                    diff = (dd - now).days
                    if diff < 0:
                        overdue.append({'project_id': p['id'], 'project_name': p['name'], 'type': 'stage', 'stage_name': s['name'], 'days_overdue': abs(diff)})
                except:
                    pass
    return jsonify({'status': 'success', 'overdue': overdue})


@app.route('/api/projects', methods=['POST'])
def api_create_project():
    """创建新项目（增强版）"""
    try:
        data = request.json
        projects = load_projects()
        import copy
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_project = {
            'id': str(int(datetime.now().timestamp() * 1000)),
            'name': data.get('name', '未命名项目'),
            'client': data.get('client', ''),
            'product_type': data.get('product_type', 'INTERNET'),
            'region': data.get('region', ''),
            'bandwidth': data.get('bandwidth', ''),
            'created_at': now_str,
            'updated_at': now_str,
            'stages': copy.deepcopy(DEFAULT_STAGES),
            'overall_progress': 0,
            'status': 'active',
            # 新增字段
            'priority': data.get('priority', 'medium'),
            'team': data.get('team', []),
            'due_date': data.get('due_date', ''),
            'comments': [],
            'activity_log': [],
            'attachments': [],
            'pricing_ref': data.get('pricing_ref', ''),
            'tags': data.get('tags', []),
        }
        add_activity_log(new_project, '创建项目', f'项目 "{new_project["name"]}" 已创建')
        projects.append(new_project)
        save_projects(projects)
        add_notification(new_project['id'], new_project['name'], 'info', f'新项目已创建: {new_project["name"]}')
        print(f"[项目监察] 新建项目: {new_project['name']}")
        return jsonify({'status': 'success', 'project': new_project})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/projects/<project_id>', methods=['PUT'])
def api_update_project(project_id):
    """更新项目（增强版 - 含工作流验证、自动日志、通知）"""
    try:
        data = request.json
        projects = load_projects()
        project = next((p for p in projects if p['id'] == project_id), None)
        if not project:
            return jsonify({'status': 'error', 'message': '项目不存在'})

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 更新阶段（含工作流验证）
        if 'stages' in data:
            old_stages = {s['id']: s for s in project.get('stages', [])}
            new_stages = data['stages']
            for ns in new_stages:
                sid = ns.get('id')
                old = old_stages.get(sid, {})
                old_status = old.get('status', 'pending')
                new_status = ns.get('status', 'pending')
                old_progress = old.get('progress', 0)
                new_progress = ns.get('progress', 0)

                # 工作流验证：检查前置阶段是否完成
                blocked_by = ns.get('blocked_by', old.get('blocked_by', []))
                if new_status != 'pending' and old_status == 'pending' and blocked_by:
                    for dep_id in blocked_by:
                        dep_stage = next((s for s in new_stages if s.get('id') == dep_id), None)
                        if dep_stage and dep_stage.get('progress', 0) < 100:
                            dep_name = dep_stage.get('name', f'阶段{dep_id}')
                            return jsonify({'status': 'error', 'message': f'无法开始此阶段：前置阶段 "{dep_name}" 尚未完成'})

                # 自动记录 started_at
                if old_status == 'pending' and new_status != 'pending':
                    if not ns.get('started_at'):
                        ns['started_at'] = now_str
                    add_activity_log(project, '阶段开始', f'阶段 "{ns.get("name", "")}" 已开始')

                # 自动记录 completed_at
                if new_progress >= 100 and old_progress < 100:
                    if not ns.get('completed_at'):
                        ns['completed_at'] = now_str
                    add_activity_log(project, '阶段完成', f'阶段 "{ns.get("name", "")}" 已完成')

                # 状态变更日志
                if old_status != new_status:
                    status_names = {'pending': '待开始', 'green': '顺利', 'yellow': '有问题', 'red': '严重问题'}
                    add_activity_log(project, '状态变更', f'阶段 "{ns.get("name", "")}" 状态: {status_names.get(old_status, old_status)} → {status_names.get(new_status, new_status)}')

                    # 触发通知
                    if new_status == 'red':
                        add_notification(project_id, project['name'], 'critical', f'阶段 "{ns.get("name", "")}" 出现严重问题')
                    elif new_status == 'yellow':
                        add_notification(project_id, project['name'], 'warning', f'阶段 "{ns.get("name", "")}" 出现问题')

            project['stages'] = new_stages

        # 更新基本信息
        for field in ['name', 'client', 'status', 'priority', 'due_date', 'team', 'tags', 'pricing_ref']:
            if field in data:
                old_val = project.get(field)
                project[field] = data[field]
                if field == 'status' and old_val != data[field]:
                    add_activity_log(project, '项目状态', f'项目状态变更为: {data[field]}')
                    if data[field] == 'completed':
                        add_notification(project_id, project['name'], 'complete', f'项目 "{project["name"]}" 已完成')

        # 重新计算总进度
        stages = project.get('stages', [])
        if stages:
            total_progress = sum(s.get('progress', 0) for s in stages)
            project['overall_progress'] = round(total_progress / len(stages))
        project['updated_at'] = now_str

        save_projects(projects)
        print(f"[项目监察] 更新项目: {project['name']} -> {project['overall_progress']}%")
        return jsonify({'status': 'success', 'project': project})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/projects/<project_id>', methods=['DELETE'])
def api_delete_project(project_id):
    """删除项目"""
    try:
        projects = load_projects()
        projects = [p for p in projects if p['id'] != project_id]
        save_projects(projects)
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


# --- 评论 API ---
@app.route('/api/projects/<project_id>/comments', methods=['POST'])
def api_add_comment(project_id):
    """添加项目评论"""
    try:
        data = request.json
        projects = load_projects()
        project = next((p for p in projects if p['id'] == project_id), None)
        if not project:
            return jsonify({'status': 'error', 'message': '项目不存在'})

        comment = {
            'id': str(int(datetime.now().timestamp() * 1000)),
            'author': data.get('author', '管理员'),
            'text': data.get('text', ''),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'type': 'comment'
        }
        if 'comments' not in project:
            project['comments'] = []
        project['comments'].insert(0, comment)
        add_activity_log(project, '新评论', f'{comment["author"]}: {comment["text"][:50]}')
        project['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_projects(projects)
        return jsonify({'status': 'success', 'comment': comment})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/projects/<project_id>/activity', methods=['GET'])
def api_get_activity(project_id):
    """获取项目操作日志"""
    projects = load_projects()
    project = next((p for p in projects if p['id'] == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'})
    return jsonify({'status': 'success', 'activity_log': project.get('activity_log', []), 'comments': project.get('comments', [])})


# --- 文件附件 API ---
@app.route('/api/projects/<project_id>/attachments', methods=['POST'])
def api_upload_attachment(project_id):
    """上传文件附件"""
    try:
        projects = load_projects()
        project = next((p for p in projects if p['id'] == project_id), None)
        if not project:
            return jsonify({'status': 'error', 'message': '项目不存在'})

        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': '没有上传文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': '文件名为空'})

        # 检查文件扩展名
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({'status': 'error', 'message': f'不支持的文件类型: .{ext}'})

        # 检查文件大小
        file.seek(0, 2)
        size = file.tell()
        file.seek(0)
        if size > MAX_FILE_SIZE:
            return jsonify({'status': 'error', 'message': f'文件超过10MB限制'})

        # 保存文件
        file_id = str(int(datetime.now().timestamp() * 1000))
        safe_name = f"{file_id}_{file.filename}"
        project_dir = os.path.join(UPLOADS_DIR, project_id)
        os.makedirs(project_dir, exist_ok=True)
        filepath = os.path.join(project_dir, safe_name)
        file.save(filepath)

        stage_id = request.form.get('stage_id', 0)
        attachment = {
            'id': file_id,
            'filename': file.filename,
            'saved_as': safe_name,
            'stage_id': int(stage_id) if stage_id else 0,
            'uploaded_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'size': size
        }
        if 'attachments' not in project:
            project['attachments'] = []
        project['attachments'].append(attachment)
        add_activity_log(project, '上传文件', f'文件 "{file.filename}" 已上传')
        project['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_projects(projects)
        return jsonify({'status': 'success', 'attachment': attachment})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/projects/<project_id>/attachments/<file_id>', methods=['GET'])
def api_download_attachment(project_id, file_id):
    """下载文件附件"""
    from flask import send_file
    try:
        projects = load_projects()
        project = next((p for p in projects if p['id'] == project_id), None)
        if not project:
            return jsonify({'status': 'error', 'message': '项目不存在'})

        attachment = next((a for a in project.get('attachments', []) if a['id'] == file_id), None)
        if not attachment:
            return jsonify({'status': 'error', 'message': '附件不存在'})

        filepath = os.path.join(UPLOADS_DIR, project_id, attachment['saved_as'])
        if not os.path.exists(filepath):
            return jsonify({'status': 'error', 'message': '文件已丢失'})

        return send_file(filepath, as_attachment=True, download_name=attachment['filename'])
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/projects/<project_id>/attachments/<file_id>', methods=['DELETE'])
def api_delete_attachment(project_id, file_id):
    """删除文件附件"""
    try:
        projects = load_projects()
        project = next((p for p in projects if p['id'] == project_id), None)
        if not project:
            return jsonify({'status': 'error', 'message': '项目不存在'})

        attachment = next((a for a in project.get('attachments', []) if a['id'] == file_id), None)
        if not attachment:
            return jsonify({'status': 'error', 'message': '附件不存在'})

        # 删除文件
        filepath = os.path.join(UPLOADS_DIR, project_id, attachment['saved_as'])
        if os.path.exists(filepath):
            os.remove(filepath)

        project['attachments'] = [a for a in project['attachments'] if a['id'] != file_id]
        add_activity_log(project, '删除文件', f'文件 "{attachment["filename"]}" 已删除')
        project['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_projects(projects)
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


# --- 通知 API ---
@app.route('/api/notifications', methods=['GET'])
def api_get_notifications():
    """获取通知列表"""
    try:
        if not os.path.exists(NOTIFICATIONS_PATH):
            return jsonify({'status': 'success', 'notifications': []})
        with open(NOTIFICATIONS_PATH, 'r', encoding='utf-8') as f:
            notifications = json.load(f)
        return jsonify({'status': 'success', 'notifications': notifications})
    except Exception as e:
        return jsonify({'status': 'success', 'notifications': []})


@app.route('/api/notifications/read', methods=['POST'])
def api_mark_notifications_read():
    """标记通知为已读"""
    try:
        if not os.path.exists(NOTIFICATIONS_PATH):
            return jsonify({'status': 'success'})
        with open(NOTIFICATIONS_PATH, 'r', encoding='utf-8') as f:
            notifications = json.load(f)
        for n in notifications:
            n['read'] = True
        with open(NOTIFICATIONS_PATH, 'w', encoding='utf-8') as f:
            json.dump(notifications, f, ensure_ascii=False, indent=2)
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


# --- 导出 API ---
@app.route('/api/projects/export/csv', methods=['GET'])
def api_export_csv():
    """导出所有项目为CSV"""
    from flask import Response
    import io
    projects = load_projects()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['项目名称', '客户', '产品类型', '区域', '带宽(Mbps)', '优先级', '截止日期', '状态', '总进度', '创建时间', '更新时间'])
    for p in projects:
        writer.writerow([
            p.get('name', ''), p.get('client', ''), p.get('product_type', ''),
            p.get('region', ''), p.get('bandwidth', ''), p.get('priority', ''),
            p.get('due_date', ''), p.get('status', ''), p.get('overall_progress', 0),
            p.get('created_at', ''), p.get('updated_at', '')
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=projects_export.csv'}
    )


@app.route('/api/projects/<project_id>/export', methods=['GET'])
def api_export_project(project_id):
    """导出单个项目详情为可打印HTML"""
    projects = load_projects()
    project = next((p for p in projects if p['id'] == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'})
    return render_template('project_print.html', project=project)


@app.route('/api/projects/<project_id>/export/pdf', methods=['GET'])
def api_export_project_pdf(project_id):
    """导出单个项目报告为 PDF"""
    from flask import Response
    import io
    projects = load_projects()
    project = next((p for p in projects if p['id'] == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'})

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=25*mm, bottomMargin=20*mm, leftMargin=20*mm, rightMargin=20*mm)

        # Try to register a CJK font for Chinese support
        font_name = 'Helvetica'
        bold_font = 'Helvetica-Bold'
        try:
            import platform
            system = platform.system()
            cjk_font_paths = []
            if system == 'Darwin':
                cjk_font_paths = [
                    '/System/Library/Fonts/STHeiti Light.ttc',
                    '/System/Library/Fonts/PingFang.ttc',
                    '/System/Library/Fonts/Hiragino Sans GB.ttc',
                    '/Library/Fonts/Arial Unicode.ttf',
                ]
            elif system == 'Windows':
                cjk_font_paths = ['C:/Windows/Fonts/msyh.ttc', 'C:/Windows/Fonts/simsun.ttc']
            else:
                cjk_font_paths = ['/usr/share/fonts/truetype/wqy/wqy-microhei.ttc', '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc']

            for fp in cjk_font_paths:
                if os.path.exists(fp):
                    pdfmetrics.registerFont(TTFont('CJK', fp))
                    font_name = 'CJK'
                    bold_font = 'CJK'
                    break
        except Exception:
            pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontName=bold_font, fontSize=18, textColor=colors.HexColor('#E60012'), spaceAfter=6)
        subtitle_style = ParagraphStyle('CustomSub', parent=styles['Normal'], fontName=font_name, fontSize=10, textColor=colors.gray, spaceAfter=14)
        section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontName=bold_font, fontSize=13, textColor=colors.HexColor('#333333'), spaceBefore=14, spaceAfter=8)
        body_style = ParagraphStyle('CustomBody', parent=styles['Normal'], fontName=font_name, fontSize=10, leading=14)
        small_style = ParagraphStyle('Small', parent=styles['Normal'], fontName=font_name, fontSize=9, textColor=colors.HexColor('#666666'))

        elements = []

        # Title
        elements.append(Paragraph(f"{project['name']} - 项目进度报告", title_style))
        elements.append(Paragraph(f"澳洲联通 项目进度管理系统 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E60012'), spaceAfter=10))

        # Project Info
        elements.append(Paragraph("项目信息", section_style))
        status_names = {'active': '进行中', 'completed': '已完成', 'cancelled': '已取消'}
        priority_names = {'high': '高', 'medium': '中', 'low': '低'}
        info_data = [
            ['客户', project.get('client', '未指定'), '产品类型', project.get('product_type', '-')],
            ['区域邮编', project.get('region', '-'), '带宽', f"{project.get('bandwidth', '-')} Mbps"],
            ['优先级', priority_names.get(project.get('priority', 'medium'), '中'), '状态', status_names.get(project.get('status', ''), project.get('status', ''))],
            ['创建时间', project.get('created_at', '-'), '更新时间', project.get('updated_at', '-')],
            ['截止日期', project.get('due_date', '未设置') or '未设置', '', ''],
        ]
        info_table = Table(info_data, colWidths=[80, 150, 80, 150])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#888888')),
            ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#888888')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        # Stage Details
        elements.append(Paragraph("阶段详情", section_style))
        status_display = {'pending': '待开始', 'green': '顺利', 'yellow': '有问题', 'red': '严重'}
        stage_header = ['#', '阶段', '状态', '完成情况', '负责人', '起始日期', '截止日期', '备注']
        stage_data = [stage_header]
        for s in project.get('stages', []):
            completion = '已完成' if s.get('progress', 0) >= 100 else '暂未完成'
            stage_data.append([
                str(s.get('id', '')),
                s.get('name', ''),
                status_display.get(s.get('status', 'pending'), s.get('status', '')),
                completion,
                s.get('assignee', '-') or '-',
                s.get('start_date', '-') or '-',
                s.get('due_date', '-') or '-',
                (s.get('notes', '') or '')[:40],
            ])
        stage_table = Table(stage_data, colWidths=[22, 62, 48, 52, 52, 58, 58, 108])
        stage_colors = {'待开始': colors.HexColor('#999999'), '顺利': colors.HexColor('#30d158'), '有问题': colors.HexColor('#ff9f0a'), '严重': colors.HexColor('#ff453a')}
        style_cmds = [
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ]
        for row_i, row in enumerate(stage_data[1:], 1):
            st_color = stage_colors.get(row[2])
            if st_color:
                style_cmds.append(('TEXTCOLOR', (2, row_i), (2, row_i), st_color))
        stage_table.setStyle(TableStyle(style_cmds))
        elements.append(stage_table)
        elements.append(Spacer(1, 14))

        # Team
        team = project.get('team', [])
        if team:
            elements.append(Paragraph("团队成员", section_style))
            team_names = ', '.join([m['name'] if isinstance(m, dict) else m for m in team])
            elements.append(Paragraph(team_names, body_style))
            elements.append(Spacer(1, 10))

        # Activity Log (last 10)
        logs = project.get('activity_log', [])[:10]
        if logs:
            elements.append(Paragraph("操作日志", section_style))
            log_data = [['时间', '操作', '详情']]
            for log in logs:
                log_data.append([log.get('timestamp', ''), log.get('action', ''), (log.get('detail', ''))[:60]])
            log_table = Table(log_data, colWidths=[120, 80, 260])
            log_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), bold_font),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(log_table)

        # Footer
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=6))
        elements.append(Paragraph("澳洲联通 | 1300-UNICOM | 24小时中英双语客服 | 99.99% SLA", small_style))
        elements.append(Paragraph(f"&copy; {datetime.now().year} 澳洲联通 Australia Unicom", small_style))

        doc.build(elements)
        buf.seek(0)
        safe_name = project['id']
        return Response(
            buf.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename=project_report_{safe_name}.pdf'}
        )
    except ImportError:
        return jsonify({'status': 'error', 'message': 'PDF export requires reportlab. Run: pip install reportlab'})
    except Exception as e:
        print(f"[PDF Export] Error: {e}")
        return jsonify({'status': 'error', 'message': f'PDF generation failed: {str(e)}'})


@app.route('/api/projects/<project_id>/export/excel', methods=['GET'])
def api_export_project_excel(project_id):
    """导出单个项目报告为 Excel"""
    from flask import Response
    import io
    projects = load_projects()
    project = next((p for p in projects if p['id'] == project_id), None)
    if not project:
        return jsonify({'status': 'error', 'message': '项目不存在'})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # --- Sheet 1: Project Overview ---
        ws = wb.active
        ws.title = "项目概览"

        # Styles
        title_font = Font(name='Arial', size=16, bold=True, color='E60012')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='E60012', end_color='E60012', fill_type='solid')
        label_font = Font(name='Arial', size=10, color='888888')
        value_font = Font(name='Arial', size=10, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{project['name']} - 项目进度报告"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:H2')
        ws['A2'] = f"澳洲联通 项目进度管理系统 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(name='Arial', size=9, color='888888')
        ws.row_dimensions[2].height = 18

        # Project Info section
        row = 4
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = '项目信息'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='333333')
        row += 1

        status_names = {'active': '进行中', 'completed': '已完成', 'cancelled': '已取消'}
        priority_names = {'high': '高', 'medium': '中', 'low': '低'}
        info_pairs = [
            ('客户', project.get('client', '未指定'), '产品类型', project.get('product_type', '-')),
            ('区域邮编', project.get('region', '-'), '带宽', f"{project.get('bandwidth', '-')} Mbps"),
            ('优先级', priority_names.get(project.get('priority', 'medium'), '中'), '状态', status_names.get(project.get('status', ''), project.get('status', ''))),
            ('创建时间', project.get('created_at', '-'), '更新时间', project.get('updated_at', '-')),
            ('截止日期', project.get('due_date', '未设置') or '未设置', '', ''),
        ]
        for pair in info_pairs:
            ws[f'A{row}'] = pair[0]
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = pair[1]
            ws[f'B{row}'].font = value_font
            ws[f'D{row}'] = pair[2]
            ws[f'D{row}'].font = label_font
            ws[f'E{row}'] = pair[3]
            ws[f'E{row}'].font = value_font
            row += 1

        # Stage Details
        row += 1
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = '阶段详情'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='333333')
        row += 1

        stage_headers = ['#', '阶段', '状态', '完成情况', '负责人', '起始日期', '截止日期', '备注']
        for col_i, h in enumerate(stage_headers, 1):
            cell = ws.cell(row=row, column=col_i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1

        status_display = {'pending': '待开始', 'green': '顺利', 'yellow': '有问题', 'red': '严重'}
        status_colors = {'pending': '999999', 'green': '30d158', 'yellow': 'ff9f0a', 'red': 'ff453a'}
        for s in project.get('stages', []):
            st = s.get('status', 'pending')
            completion = '已完成' if s.get('progress', 0) >= 100 else '暂未完成'
            vals = [s.get('id', ''), s.get('name', ''), status_display.get(st, st), completion, s.get('assignee', '-') or '-', s.get('start_date', '-') or '-', s.get('due_date', '-') or '-', s.get('notes', '') or '']
            for col_i, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_i, value=v)
                cell.font = Font(name='Arial', size=10)
                cell.border = thin_border
                if col_i == 3:
                    cell.font = Font(name='Arial', size=10, bold=True, color=status_colors.get(st, '333333'))
                if col_i in (1, 4):
                    cell.alignment = Alignment(horizontal='center')
            row += 1

        # Team
        team = project.get('team', [])
        if team:
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = '团队成员'
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='333333')
            row += 1
            team_names = ', '.join([m['name'] if isinstance(m, dict) else m for m in team])
            ws[f'A{row}'] = team_names
            ws[f'A{row}'].font = Font(name='Arial', size=10)
            row += 1

        # Column widths
        col_widths = [6, 16, 12, 12, 14, 14, 14, 26]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # --- Sheet 2: Activity Log ---
        ws2 = wb.create_sheet('操作日志')
        logs = project.get('activity_log', [])
        ws2.merge_cells('A1:C1')
        ws2['A1'] = f"{project['name']} - 操作日志"
        ws2['A1'].font = Font(name='Arial', size=14, bold=True, color='E60012')

        log_headers = ['时间', '操作', '详情']
        for col_i, h in enumerate(log_headers, 1):
            cell = ws2.cell(row=3, column=col_i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, log in enumerate(logs, 4):
            ws2.cell(row=i, column=1, value=log.get('timestamp', '')).border = thin_border
            ws2.cell(row=i, column=2, value=log.get('action', '')).border = thin_border
            ws2.cell(row=i, column=3, value=log.get('detail', '')).border = thin_border

        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 50

        # Footer row on overview sheet
        row += 2
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f'澳洲联通 | 1300-UNICOM | © {datetime.now().year} 澳洲联通 Australia Unicom'
        ws[f'A{row}'].font = Font(name='Arial', size=8, color='AAAAAA')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = project['id']
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename=project_report_{safe_name}.xlsx'}
        )
    except ImportError:
        return jsonify({'status': 'error', 'message': 'Excel export requires openpyxl. Run: pip install openpyxl'})
    except Exception as e:
        print(f"[Excel Export] Error: {e}")
        return jsonify({'status': 'error', 'message': f'Excel generation failed: {str(e)}'})


# --- 4. 预测接口 ---
@app.route('/predict', methods=['POST'])
@limiter.limit("30 per minute")
def predict():
    try:
        data = request.json
        print(f"收到请求: {data}")

        # 1. 提取并验证用户输入的邮编
        user_region = str(data['region'])
        is_valid, region_name = validate_postcode(user_region)
        if not is_valid:
            return jsonify({
                'status': 'error',
                'message': f'邮编无效：{region_name}。请输入有效的澳洲邮编（如 2000、3000、4000）。'
            })

        # 2. 翻译成 Zone (这是新模型需要的特征)
        derived_zone = get_zone_from_postcode(user_region)
        print(f"[地区转换] 用户输入 '{user_region}' ({region_name}) -> AI识别为 '{derived_zone}'")

        # 3. 整理预测数据 (列名顺序要和训练时一致)
        X_pred = pd.DataFrame([{
            'operator': data['operator'],
            'product_type': data['product_type'],
            'bandwidth_mbps': float(data['bandwidth']),
            'term_months': int(data['term']),
            'zone': derived_zone  # 注意：这里传的是 zone，不是 region
        }])


        # 4. 预测
        raw_price = model.predict(X_pred)[0]

        # === [新增] 业务保底逻辑 (Business Rules) ===
        # 即使 AI 算出了负数，我们也不能展示给客户看
        MIN_PRICE = 50.0  # 设定：全网最低起步价 (你可以改成 100 或其他数字)

        if raw_price < MIN_PRICE:
            print(f"[警告] 触发保底: AI 算出 {raw_price}，已修正为 {MIN_PRICE}")
            price = MIN_PRICE
        else:
            price = raw_price

        # === [改进] 基于实际数据匹配度的置信度计算 ===
        confidence = 0
        
        if df_training is not None:
            # 方法1: 精确匹配 - 查找完全相同的参数组合
            exact_match = df_training[
                (df_training['operator'] == data['operator']) &
                (df_training['product_type'] == data['product_type']) &
                (df_training['zone'] == derived_zone) &
                (df_training['term_months'] == int(data['term']))
            ]
            
            if len(exact_match) > 0:
                # 有完全匹配的记录
                bandwidth = float(data['bandwidth'])
                # 检查带宽是否接近训练数据中的值
                bandwidth_diff = abs(exact_match['bandwidth_mbps'] - bandwidth)
                min_diff = bandwidth_diff.min()
                
                if min_diff <= 10:  # 带宽差异小于10Mbps
                    confidence = 85  # 高置信度
                elif min_diff <= 50:  # 带宽差异小于50Mbps
                    confidence = 70  # 中等置信度
                else:
                    confidence = 55  # 较低置信度
            else:
                # 方法2: 部分匹配 - 检查各个维度的数据覆盖度
                operator_count = len(df_training[df_training['operator'] == data['operator']])
                product_count = len(df_training[df_training['product_type'] == data['product_type']])
                zone_count = len(df_training[df_training['zone'] == derived_zone])
                term_count = len(df_training[df_training['term_months'] == int(data['term'])])
                
                total_records = len(df_training)
                
                # 计算各维度的覆盖率
                operator_coverage = operator_count / total_records * 100
                product_coverage = product_count / total_records * 100
                zone_coverage = zone_count / total_records * 100
                term_coverage = term_count / total_records * 100
                
                # 检查带宽范围
                bandwidth = float(data['bandwidth'])
                bandwidth_in_range = (df_training['bandwidth_mbps'].min() <= bandwidth <= df_training['bandwidth_mbps'].max())
                
                # 综合计算置信度（加权平均）
                # 运营商和产品类型权重较高
                base_confidence = (operator_coverage * 0.3 + product_coverage * 0.3 + 
                                  zone_coverage * 0.2 + term_coverage * 0.1)
                
                # 如果带宽在范围内，增加置信度
                if bandwidth_in_range:
                    base_confidence += 5
                
                # 由于数据量小（286条），整体降低置信度基准
                confidence = base_confidence * 0.85  # 降低15%以反映数据稀疏性
                
                # 如果某个维度数据很少，进一步降低
                if operator_count < 5 or product_count < 5 or zone_count < 3:
                    confidence *= 0.8  # 再降低20%
                
                # 确保置信度在合理范围内
                confidence = max(25, min(75, confidence))  # 限制在25-75之间（无精确匹配时）
        else:
            # 如果无法加载训练数据，使用简化算法
            confidence = 50  # 默认中等置信度
        
        # 最终限制在 0-100 范围内
        confidence = max(0, min(100, round(confidence, 0)))

        # 5. 构建参考依据（evidence）- 从训练数据中找相似记录
        evidence = []
        if df_training is not None:
            # 查找同运营商、同产品类型的记录，按带宽差异排序
            similar = df_training[
                (df_training['operator'] == data['operator']) &
                (df_training['product_type'] == data['product_type'])
            ].copy()
            if len(similar) == 0:
                # 如果没有完全匹配，放宽条件只匹配产品类型
                similar = df_training[
                    df_training['product_type'] == data['product_type']
                ].copy()
            if len(similar) > 0:
                similar['_bw_diff'] = abs(similar['bandwidth_mbps'] - float(data['bandwidth']))
                similar = similar.sort_values('_bw_diff').head(3)
                for _, row in similar.iterrows():
                    evidence.append({
                        'operator': row['operator'],
                        'bandwidth': int(row['bandwidth_mbps']),
                        'price': round(float(row['price']), 2)
                    })

        # 6. 返回结果
        return jsonify({
            'status': 'success',
            'price': round(price, 2),
            'confidence': round(confidence, 0),
            'zone': derived_zone,
            'evidence': evidence,
            'debug_info': f"已按 [{derived_zone}] 区域标准计算"
        })

    except Exception as e:
        print(f"预测出错: {e}")
        return jsonify({'status': 'error', 'message': str(e)})


# --- 人工核价接口 ---
MANUAL_REVIEW_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "manual_reviews.csv")

@app.route('/manual-review', methods=['POST'])
def manual_review():
    """接收人工核价申请，保存到 CSV 文件"""
    try:
        data = request.json
        row = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'operator': data.get('operator', ''),
            'product_type': data.get('product_type', ''),
            'bandwidth': data.get('bandwidth', ''),
            'region': data.get('region', ''),
            'term': data.get('term', ''),
            'ai_price': data.get('ai_price', ''),
            'confidence': data.get('confidence', ''),
        }
        file_exists = os.path.exists(MANUAL_REVIEW_PATH)
        with open(MANUAL_REVIEW_PATH, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=row.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)
        print(f"[人工核价] 已保存申请: {row['operator']} {row['bandwidth']}Mbps {row['region']}")
        return jsonify({'status': 'success', 'message': '核价申请已提交'})
    except Exception as e:
        print(f"[人工核价] 保存失败: {e}")
        return jsonify({'status': 'error', 'message': str(e)})


# --- 用户反馈接口 ---
FEEDBACK_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "user_feedback.csv")

@app.route('/feedback', methods=['POST'])
def feedback():
    """接收用户对估价结果的反馈"""
    try:
        data = request.json
        row = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'operator': data.get('operator', ''),
            'product_type': data.get('product_type', ''),
            'bandwidth': data.get('bandwidth', ''),
            'region': data.get('region', ''),
            'term': data.get('term', ''),
            'ai_price': data.get('ai_price', ''),
            'confidence': data.get('confidence', ''),
            'rating': data.get('rating', ''),  # 'accurate' 或 'inaccurate'
        }
        file_exists = os.path.exists(FEEDBACK_PATH)
        with open(FEEDBACK_PATH, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=row.keys())
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)
        print(f"[用户反馈] {row['rating']}: {row['operator']} {row['bandwidth']}Mbps -> ${row['ai_price']}")
        return jsonify({'status': 'success', 'message': '感谢您的反馈'})
    except Exception as e:
        print(f"[用户反馈] 保存失败: {e}")
        return jsonify({'status': 'error', 'message': str(e)})


# --- AI 模型调用函数 ---
def call_lmstudio(system_prompt: str, user_message: str) -> str:
    """调用 LM Studio 本地模型（完全免费）"""
    try:
        # 先检查服务器是否运行
        try:
            check_response = requests.get(f"{LMSTUDIO_BASE_URL}/v1/models", timeout=3)
            if check_response.status_code != 200:
                raise Exception("LM Studio 服务器未正常响应")
        except requests.exceptions.ConnectionError:
            raise Exception("无法连接到 LM Studio 服务器。\n\n请确保：\n1. LM Studio 已打开\n2. 模型已加载（Qwen3-VL-8B-Instruct）\n3. 本地服务器已启动（点击右下角 'Local Server' 或 'Server' 按钮）\n4. 服务器运行在端口 1234")
        except requests.exceptions.Timeout:
            raise Exception("LM Studio 服务器响应超时。\n\n可能原因：\n1. 服务器正在启动中，请等待 30 秒后重试\n2. 模型正在加载，请等待加载完成\n3. 服务器负载过高")
        
        # LM Studio 使用 OpenAI 兼容的 API 格式
        payload = {
            'model': LMSTUDIO_MODEL if LMSTUDIO_MODEL else 'local-model',  # 如果未指定，使用默认
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_message}
            ],
            'temperature': 0.7,
            'max_tokens': 1500,  # 降低 token 数量以加快生成速度
            'stream': False
        }
        
        url = f"{LMSTUDIO_BASE_URL}/v1/chat/completions"
        # 增加超时时间到 180 秒（8B 模型在 CPU 模式下可能需要 2-3 分钟）
        response = requests.post(url, json=payload, timeout=180)
        
        if response.status_code == 200:
            result = response.json()
            if 'choices' in result and len(result['choices']) > 0:
                return result['choices'][0]['message']['content']
            else:
                raise Exception("LM Studio 返回格式异常")
        else:
            raise Exception(f"LM Studio API 错误: {response.status_code} - {response.text}")
    except requests.exceptions.Timeout:
        raise Exception("LM Studio 响应超时（180秒）。\n\n可能原因：\n1. 模型生成时间较长（8B 模型在 CPU 模式下需要 2-3 分钟）\n2. 服务器负载过高\n3. 未启用 GPU 加速\n\n建议：\n1. 在 LM Studio 设置中启用 GPU 加速（你有 RTX 3070 Ti）\n2. 检查 LM Studio 界面，确认模型已完全加载\n3. 尝试重启 LM Studio 服务器\n4. 如果持续超时，考虑使用更小的模型（7B Q4 量化版）")
    except requests.exceptions.ConnectionError:
        raise Exception("无法连接到 LM Studio 服务器。\n\n请确保：\n1. LM Studio 已打开\n2. 模型已加载（Qwen3-VL-8B-Instruct）\n3. 本地服务器已启动（点击右下角 'Local Server' 或 'Server' 按钮）\n4. 服务器运行在端口 1234")
    except Exception as e:
        error_msg = str(e)
        if "timeout" in error_msg.lower():
            raise Exception(f"LM Studio 响应超时。\n\n请检查：\n1. LM Studio 服务器是否正常运行\n2. 模型是否已完全加载\n3. 可以尝试重启 LM Studio 服务器")
        else:
            raise Exception(f"LM Studio 调用失败: {error_msg}")

def call_ollama(system_prompt: str, user_message: str) -> str:
    """调用 Ollama 本地模型（完全免费）"""
    try:
        payload = {
            'model': OLLAMA_MODEL,
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_message}
            ],
            'stream': False
        }
        response = requests.post(f"{OLLAMA_BASE_URL}/api/chat", json=payload, timeout=180)
        if response.status_code == 200:
            return response.json()['message']['content']
        else:
            raise Exception(f"Ollama API 错误: {response.status_code}")
    except requests.exceptions.ConnectionError:
        raise Exception("无法连接到 Ollama。请确保已安装并启动 Ollama 服务。\n安装: https://ollama.com/download\n启动后运行: ollama pull qwen2.5:7b")
    except Exception as e:
        raise Exception(f"Ollama 调用失败: {str(e)}")

def call_gemini(system_prompt: str, user_message: str) -> str:
    """调用 Google Gemini API（使用 gemini-2.0-flash 模型）"""
    if not GEMINI_API_KEY:
        raise Exception("Gemini API Key 未配置。请设置环境变量 GEMINI_API_KEY")
    
    try:
        # 构建完整的提示词
        full_prompt = f"{system_prompt}\n\n用户问题: {user_message}"
        
        # 使用 X-goog-api-key header 格式（根据你提供的 curl 示例）
        headers = {
            'Content-Type': 'application/json',
            'X-goog-api-key': GEMINI_API_KEY
        }
        
        payload = {
            'contents': [{
                'parts': [{
                    'text': full_prompt
                }]
            }]
        }
        
        # 使用 gemini-2.0-flash 模型
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
        
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            # 提取回复文本
            if 'candidates' in result and len(result['candidates']) > 0:
                return result['candidates'][0]['content']['parts'][0]['text']
            else:
                raise Exception("Gemini API 返回格式异常")
        else:
            error_text = response.text
            raise Exception(f"Gemini API 错误: {response.status_code} - {error_text}")
    except Exception as e:
        raise Exception(f"Gemini 调用失败: {str(e)}")

def call_qwen(system_prompt: str, user_message: str) -> str:
    """调用通义千问 API（国内访问快，有免费额度）"""
    if not QWEN_API_KEY:
        raise Exception("通义千问 API Key 未配置。请设置环境变量 QWEN_API_KEY")
    
    try:
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {QWEN_API_KEY}'
        }
        payload = {
            'model': 'qwen-turbo',
            'input': {
                'messages': [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_message}
                ]
            },
            'parameters': {
                'temperature': 0.7,
                'max_tokens': 2000
            }
        }
        response = requests.post(QWEN_API_URL, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            return response.json()['output']['text']
        else:
            raise Exception(f"通义千问 API 错误: {response.status_code}")
    except Exception as e:
        raise Exception(f"通义千问调用失败: {str(e)}")

def call_deepseek(system_prompt: str, user_message: str) -> str:
    """调用 DeepSeek API（备用）"""
    if not DEEPSEEK_API_KEY:
        raise Exception("DeepSeek API Key 未配置。请设置环境变量 DEEPSEEK_API_KEY")
    
    try:
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {DEEPSEEK_API_KEY}'
        }
        payload = {
            'model': 'deepseek-chat',
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': user_message}
            ],
            'temperature': 0.7,
            'max_tokens': 2000
        }
        response = requests.post(DEEPSEEK_API_URL, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            return response.json()['choices'][0]['message']['content']
        else:
            raise Exception(f"DeepSeek API 错误: {response.status_code}")
    except Exception as e:
        raise Exception(f"DeepSeek 调用失败: {str(e)}")

# --- 网络搜索功能 ---
def search_web(query: str, max_results: int = 3) -> List[Dict]:
    """使用 DuckDuckGo 搜索网络资料（完全免费，无需 API Key）"""
    if not WEB_SEARCH_AVAILABLE:
        return []
    try:
        search_results = []
        with DDGS() as ddgs:
            # 搜索相关结果
            results = list(ddgs.text(query, max_results=max_results))
            for result in results:
                search_results.append({
                    'title': result.get('title', ''),
                    'snippet': result.get('body', ''),
                    'url': result.get('href', '')
                })
        return search_results
    except Exception as e:
        print(f"[搜索] 搜索失败: {e}")
        return []

def should_search(user_message: str) -> bool:
    """判断是否需要联网搜索"""
    # 关键词列表：如果问题包含这些词，可能需要搜索
    search_keywords = [
        '最新', '最近', '现在', '当前', '今年', '明年', '未来', '趋势',
        '新闻', '更新', '变化', '涨价', '降价', '政策', '规定',
        '市场', '行业', '竞争', '对比', '比较', '资料', '信息'
    ]
    user_lower = user_message.lower()
    return any(keyword in user_lower for keyword in search_keywords)

# --- 5. AI 分析接口（支持多个免费模型 + 联网搜索）---
@app.route('/deepseek/chat', methods=['POST'])
@limiter.limit("10 per minute")
def ai_chat():
    """AI 对话接口，基于项目数据进行专有化分析，支持多个免费模型和联网搜索"""
    try:
        data = request.json
        user_message = data.get('message', '')
        requested_model = data.get('model', '').lower()  # 从前端获取请求的模型
        enable_search = data.get('enable_search', True)  # 默认启用搜索
        
        # 如果前端指定了模型，使用前端指定的；否则使用默认配置
        model_to_use = requested_model if requested_model in ['lmstudio', 'ollama', 'gemini', 'qwen', 'deepseek'] else AI_MODEL_TYPE
        
        if not user_message:
            return jsonify({'status': 'error', 'message': '消息不能为空'})
        
        # 联网搜索（如果需要）
        search_info = ""
        if enable_search and should_search(user_message):
            print(f"[搜索] 正在搜索: {user_message[:50]}...")
            search_results = search_web(user_message, max_results=3)
            if search_results:
                search_info = "\n\n=== 网络搜索结果 ===\n"
                for i, result in enumerate(search_results, 1):
                    search_info += f"\n{i}. {result['title']}\n"
                    search_info += f"   {result['snippet'][:200]}...\n"
                    search_info += f"   来源: {result['url']}\n"
                search_info += "\n请结合以上网络搜索结果和项目数据库来回答用户问题。"
                print(f"[搜索] 找到 {len(search_results)} 条结果")
            else:
                print(f"[搜索] 未找到相关结果")
        
        # === 构建系统提示词：公司知识库 + 训练数据 + 搜索结果 ===
        # 将 company_knowledge.txt 和训练数据摘要合并为完整的内部资料
        context_sections = [company_context]  # 公司知识库（来自 company_knowledge.txt）
        if training_data_summary:
            context_sections.append(training_data_summary)  # 训练数据统计摘要
        if search_info:
            context_sections.append(search_info)  # 网络搜索结果（如有）

        context_data = "\n\n".join(context_sections)

        system_prompt = (
            "你是澳洲联通 (Australia Unicom) 的 AI 客服。"
            "请根据以下内部资料回答用户问题。"
            "如果用户问的内容不在资料中，请礼貌地引导他们拨打客服热线 1300-UNICOM 或访问官网查询。\n\n"
            "[内部资料开始]\n"
            f"{context_data}\n"
            "[内部资料结束]\n\n"
            "回答要求：\n"
            "1. 严格基于内部资料回答，不要编造不存在的信息\n"
            "2. 涉及价格时，引用官方参考价格并说明实际价格可能浮动\n"
            "3. 涉及数据分析时，引用训练数据统计摘要中的具体数字\n"
            "4. 用中文回答，专业、简洁、易懂\n"
            "5. 如果信息不足，建议用户使用 AI 估价系统或联系 1300-UNICOM\n"
            "6. 如果引用了网络搜索结果，请注明信息来源"
        )
        
        print(f"[AI-{model_to_use.upper()}] 收到问题: {user_message[:50]}...")
        
        # 根据选择的模型类型调用相应的 API
        try:
            if model_to_use == 'lmstudio':
                ai_response = call_lmstudio(system_prompt, user_message)
            elif model_to_use == 'ollama':
                ai_response = call_ollama(system_prompt, user_message)
            elif model_to_use == 'gemini':
                ai_response = call_gemini(system_prompt, user_message)
            elif model_to_use == 'qwen':
                ai_response = call_qwen(system_prompt, user_message)
            elif model_to_use == 'deepseek':
                ai_response = call_deepseek(system_prompt, user_message)
            else:
                return jsonify({
                    'status': 'error',
                    'message': f'不支持的模型类型: {model_to_use}。支持的类型: lmstudio, ollama, gemini, qwen, deepseek'
                })
            
            print(f"[AI-{model_to_use.upper()}] 回答生成成功")
            
            return jsonify({
                'status': 'success',
                'response': ai_response
            })
            
        except Exception as api_error:
            error_msg = str(api_error)
            print(f"[AI-{model_to_use.upper()}] 错误: {error_msg}")
            
            # 如果是 Gemini 配额错误，提供友好的提示
            if model_to_use == 'gemini' and ('429' in error_msg or 'quota' in error_msg.lower() or 'RESOURCE_EXHAUSTED' in error_msg):
                friendly_msg = """Gemini API 配额已用完。建议切换到其他免费模型：

1. **Ollama（推荐）** - 完全免费，本地运行，无配额限制
   - 需要先安装 Ollama：https://ollama.com/download
   - 然后运行：ollama pull qwen2.5:7b

2. **通义千问** - 有免费额度，国内访问快
   - 需要配置 QWEN_API_KEY

请在对话框顶部的模型选择器中切换到其他模型。"""
                return jsonify({
                    'status': 'error',
                    'message': friendly_msg
                })
            
            return jsonify({
                'status': 'error',
                'message': error_msg
            })
            
    except requests.exceptions.Timeout:
        return jsonify({'status': 'error', 'message': '请求超时，请稍后重试'})
    except Exception as e:
        print(f"[AI] 出错: {e}")
        return jsonify({'status': 'error', 'message': str(e)})


if __name__ == '__main__':
    # host='0.0.0.0' 允许局域网访问，避免出现 Connection Refused
    import socket
    PORT = int(os.getenv('PORT', 8080))  # macOS 5000 端口被 AirPlay 占用，默认改为 8080
    # 获取本机IP地址
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
        s.close()
        print(f"\n{'='*50}")
        print(f"服务器已启动")
        print(f"{'='*50}")
        print(f"本地访问: http://127.0.0.1:{PORT}")
        print(f"局域网访问: http://{local_ip}:{PORT}")
        print(f"{'='*50}")
        print(f"告诉同事在浏览器中输入: http://{local_ip}:{PORT}")
        print(f"{'='*50}\n")
    except:
        print(f"\n服务器已启动，监听所有网络接口 (0.0.0.0:{PORT})")
        print("请运行 ipconfig 查看您的IP地址\n")
    
    app.run(debug=False, port=PORT, host='0.0.0.0')